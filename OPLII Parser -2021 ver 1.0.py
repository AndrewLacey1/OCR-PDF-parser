
import os
from pdf2image import convert_from_path
import PyPDF2
from openpyxl import Workbook
from openpyxl import load_workbook
import difflib
import PySimpleGUI as sg
import sys
import PyInstaller

try:
    from PIL import Image
except ImportError:
    import Image
import pytesseract

ErrorCount = 0 #variable for number of pdfs in which no valid serial number was found
DocumentCount = 0 #number of files in the selected folder
wb = Workbook() #creating a workbook output
ws = wb.active #creating a worksheet output
ws.title = "Sheet 1" #naming worksheet
d = ws.cell(row=1, column=1,value='Serial No') #naming column 1
d = ws.cell(row=1, column=2,value='File Name') #naming column 2
unwantedchars = ".:*,,;'$?|’_(){}~`@[]-#" + '"' +'“' + '»' + '-' + "'" +"," + ":" + "&"+"!" #list of unwanted characters that serial number cannot contain


def remove_vals_from_list(list,val):#removes value 'val' from the input 'list'
    return [value for value in list if value != val]


def ExtractPages(FileNameOriginal, FileNameOutput, xPageStart, xPageEnd):#used to extract a range of pages from a pdf and save as their own file
    from PyPDF2 import PdfFileReader, PdfFileWriter
    output = PdfFileWriter()
    pdfOne = read_PDF
    for i in range(xPageStart, xPageEnd):
        output.addPage(pdfOne.getPage(i))
        outputStream = open(FileNameOutput, "wb")
        output.write(outputStream)
        outputStream.close()


def ocr_core(filename):
    """
    This function will handle the core OCR processing of images.
    """
    text = pytesseract.image_to_string(Image.open(filename))  # We'll use Pillow's Image class to open the image and pytesseract to detect the string in the image
    return text


Count = 0

###################################################################
#Initialization of gui
###################################################################
sg.theme('DarkTeal10') #color of gui
layout = [[sg.Text('Fill in the fields')],
          [sg.Text('Files', size=(17, 1)), sg.Input(), sg.FolderBrowse()], #folder browser for folder with input files
          [sg.Text('Serial no. Spreadsheet', size=(17, 1)), sg.Input(), sg.FileBrowse()], #file browser for input spreadsheet with serial no.
          [sg.Text('Tesseract.exe path:',size=(17,1)),sg.Input(),sg.FileBrowse()], #path for tesseract .exe, better ways to do this, but this is easiest
          [sg.Submit(), sg.Cancel()]]

window = sg.Window('Parser', layout)

event, values = window.read()
window.close()
print(f'You clicked {event}')
print(f'You chose filenames {values[0]}, {values[1]} and {values[2]}')

###################################################################
#beginning of main body of program
###################################################################
pytesseract.pytesseract.tesseract_cmd = values[2] #Telling pytesseract where tesseract.exe is
EquipmentList = load_workbook(filename= values[1])#load in input spreadsheet
Sheet = EquipmentList.worksheets[0]#select first sheet in workbook
SerialNumberList = []#initialization
NumRows = Sheet.max_row #number of rows in input sheet
Foldername = values[0] #input folder directory
FileList = os.listdir(Foldername) #list of files in 'FolderName'
NumFiles = len(FileList) #number of files in 'FileList'

for l in range (2,NumRows): #creating list of valid serial numbers from input file, assumes serial numbers begin in column E, row 2
    CurrentSerial = Sheet['E' + str(l)].value
    if CurrentSerial != None and CurrentSerial != '':
        SerialNumberList.append(CurrentSerial)

#SerialNumberList = set(SerialNumberList)
print(SerialNumberList)
print(FileList)
print("Initialized")

print(Foldername)
for j in range (0,NumFiles):
    PDF = open(Foldername + "/" +FileList[j],'rb')#open file in read mode
    read_PDF = PyPDF2.PdfFileReader(PDF)#used in extract_pages
    pages = convert_from_path(Foldername + "\\" + FileList[j], dpi=300)#imports pages into python
    i=0
    for page in pages:#for each loop on each page in pages
        sg.OneLineProgressMeter('Parser', DocumentCount + 1, NumFiles*20, 'single')#initializing/updating progress bar
        DocumentCount = DocumentCount + 1
        page.save('Placeholder.jpg', 'JPEG')#save temp image for OCR
        i = i+1
        String = ocr_core(r'Placeholder.jpg') #ocr on temp image
        os.remove('Placeholder.jpg')
        PageWords = String.split()#split up words detected on page by spaces.
        PageWords = [word.strip(unwantedchars).upper() for word in PageWords]#remove unwanted characters from each word to lower chance of missing serial number
        Numwords = len(PageWords)

        for word in PageWords:
            if word.isalpha() == True:#if word is purely letters, remove it, not a serial number
                PageWords = remove_vals_from_list(PageWords,word)
            elif word.isdigit()==True:#if word is purely numbers, remove it, not a serial number
                PageWords = remove_vals_from_list(PageWords, word)
        PageWords = set(PageWords)
        print("\n")
        print(PageWords)
        SharedStrings = list(PageWords.intersection(SerialNumberList))

        if len(SharedStrings) == 1:#if one shared string, it is  serial number
            print(SharedStrings[0])
            print(str(i))
            d = ws.cell(row=DocumentCount + 1, column=1, value=SharedStrings[0])#save serial number in spreadsheet
            d = ws.cell(row=DocumentCount + 1, column=2, value=SharedStrings[0] + '.pdf')#save name of pdf to excel spreadsheet
            ExtractPages(Foldername+"/" + str(FileList[j]),Foldername + "/" + str(SharedStrings[0])+'.pdf',i-1,i)#save pdf with name of serial number
        elif len(SharedStrings) == 0:#if no shared strings, try to find close matches, OCR may have misread
            PageWords = list(PageWords)
            Matches = []
            for p in range (0,len(PageWords)-1):#for eac
                    Matches.append(difflib.get_close_matches(PageWords[p],SerialNumberList,n=1,cutoff=0.8))#find close matches, 0.8 is tolerance
                    if Matches[p] != []:#if match is not empty, correct match found, extract pdf
                        print(Matches[p][0])
                        print("^Alternate Match")
                        d = ws.cell(row=DocumentCount+1, column=1, value=str(Matches[p][0]))
                        d = ws.cell(row=DocumentCount + 1, column=2, value=str(Matches[p][0])+'.pdf')
                        print(Foldername)
                        print(str(FileList[j]))
                        print(Foldername +'/' + str(Matches[p][0])+'.pdf')
                        ExtractPages(Foldername + "/" + str(FileList[j]),Foldername + '/' + str(Matches[p][0]) + '.pdf', i - 1, i)
                        break
            Matches = list(filter(None, Matches))
            if len(Matches) == 0:#if no match, save document as such
                ErrorCount = ErrorCount +1
                print("No Match Found")
                d = ws.cell(row=DocumentCount + 1, column=1, value="_NoserialFound" + str(ErrorCount))
                d = ws.cell(row=DocumentCount + 1, column=2, value="_NoserialFound" + str(ErrorCount) + ".pdf")
                ExtractPages(Foldername+"/" + str(FileList[j]),Foldername + "/" +"_NoSerialno." +str(ErrorCount)+'.pdf',i-1,i)
window.close()
wb.save(Foldername + '/output spreadsheet.xlsx')

